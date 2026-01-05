import sys
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QLabel, QLineEdit, QPushButton, 
                             QTextEdit, QGroupBox, QGridLayout, QDateEdit, 
                             QTableWidget, QTableWidgetItem, QHeaderView, QTabWidget, 
                             QComboBox, QDialog, QDialogButtonBox, QSpinBox, QDoubleSpinBox)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QFont, QColor
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import calendar
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcludeMonthsDialog(QDialog):
    def __init__(self, existing_exclusions=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Exclude EMI Months")
        self.setModal(True)
        self.setMinimumWidth(400)
        
        layout = QVBoxLayout(self)
        
        # Instructions
        info_label = QLabel("Select months to exclude from automatic EMI payments:")
        info_label.setWordWrap(True)
        layout.addWidget(info_label)
        
        # Input fields group
        fields_group = QGroupBox("Exclusion Details")
        fields_layout = QGridLayout()
        
        # Month and Year selection
        fields_layout.addWidget(QLabel("Month:"), 0, 0)
        self.month_combo = QComboBox()
        self.month_combo.addItems([
            "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December"
        ])
        fields_layout.addWidget(self.month_combo, 0, 1)
        
        fields_layout.addWidget(QLabel("Year:"), 1, 0)
        self.year_spinbox = QSpinBox()
        self.year_spinbox.setRange(2020, 2050)
        self.year_spinbox.setValue(QDate.currentDate().year())
        fields_layout.addWidget(self.year_spinbox, 1, 1)
        
        fields_group.setLayout(fields_layout)
        layout.addWidget(fields_group)
        
        # Add button
        add_btn = QPushButton("Add to Exclusion List")
        add_btn.clicked.connect(self.add_exclusion)
        layout.addWidget(add_btn)
        
        # List of excluded months
        layout.addWidget(QLabel("Excluded Months:"))
        self.exclusion_list = QTextEdit()
        self.exclusion_list.setReadOnly(True)
        self.exclusion_list.setMaximumHeight(150)
        layout.addWidget(self.exclusion_list)
        
        # Initialize exclusions
        self.exclusions = existing_exclusions if existing_exclusions else []
        self.update_exclusion_display()
        
        # Remove button
        remove_btn = QPushButton("Remove Last Entry")
        remove_btn.clicked.connect(self.remove_last_exclusion)
        layout.addWidget(remove_btn)
        
        # Buttons
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | 
                                     QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def add_exclusion(self):
        month = self.month_combo.currentIndex() + 1  # 1-12
        year = self.year_spinbox.value()
        
        exclusion = {'month': month, 'year': year}
        
        # Check if already exists
        if exclusion not in self.exclusions:
            self.exclusions.append(exclusion)
            self.update_exclusion_display()
    
    def remove_last_exclusion(self):
        if self.exclusions:
            self.exclusions.pop()
            self.update_exclusion_display()
    
    def update_exclusion_display(self):
        if not self.exclusions:
            self.exclusion_list.setPlainText("No months excluded yet.")
            return
        
        month_names = ["January", "February", "March", "April", "May", "June",
                      "July", "August", "September", "October", "November", "December"]
        
        text = ""
        for exc in sorted(self.exclusions, key=lambda x: (x['year'], x['month'])):
            text += f"{month_names[exc['month']-1]} {exc['year']}\n"
        
        self.exclusion_list.setPlainText(text)
    
    def get_exclusions(self):
        return self.exclusions


class ManualEMIDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Add Manual EMI")
        self.setModal(True)
        self.setMinimumWidth(400)
        
        layout = QVBoxLayout(self)
        
        # Input fields group
        fields_group = QGroupBox("EMI Details")
        fields_layout = QGridLayout()
        
        # Amount field
        fields_layout.addWidget(QLabel("EMI Amount:"), 0, 0)
        self.amount_input = QDoubleSpinBox()
        self.amount_input.setRange(0, 99999999)
        self.amount_input.setDecimals(2)
        self.amount_input.setValue(40800)
        self.amount_input.setPrefix("₹")
        fields_layout.addWidget(self.amount_input, 0, 1)
        
        # Date field
        fields_layout.addWidget(QLabel("EMI Date:"), 1, 0)
        self.date_input = QDateEdit()
        self.date_input.setCalendarPopup(True)
        self.date_input.setDate(QDate.currentDate())
        self.date_input.setDisplayFormat("dd-MM-yyyy")
        fields_layout.addWidget(self.date_input, 1, 1)
        
        # Note field (optional)
        fields_layout.addWidget(QLabel("Note (Optional):"), 2, 0)
        self.note_input = QLineEdit()
        self.note_input.setPlaceholderText("e.g., Extra payment, One-time EMI")
        fields_layout.addWidget(self.note_input, 2, 1)
        
        fields_group.setLayout(fields_layout)
        layout.addWidget(fields_group)
        
        # Buttons
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | 
                                     QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def get_emi_data(self):
        date = self.date_input.date()
        return {
            'amount': self.amount_input.value(),
            'date': datetime(date.year(), date.month(), date.day()),
            'note': self.note_input.text()
        }


class BankChargeDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Add Bank Charge")
        self.setModal(True)
        self.setMinimumWidth(400)
        
        layout = QVBoxLayout(self)
        
        # Input fields group
        fields_group = QGroupBox("Charge Details")
        fields_layout = QGridLayout()
        
        # Amount field
        fields_layout.addWidget(QLabel("Charge Amount:"), 0, 0)
        self.amount_input = QDoubleSpinBox()
        self.amount_input.setRange(0, 99999999)
        self.amount_input.setDecimals(2)
        self.amount_input.setValue(500)
        self.amount_input.setPrefix("₹")
        fields_layout.addWidget(self.amount_input, 0, 1)
        
        # Date field
        fields_layout.addWidget(QLabel("Charge Date:"), 1, 0)
        self.date_input = QDateEdit()
        self.date_input.setCalendarPopup(True)
        self.date_input.setDate(QDate.currentDate())
        self.date_input.setDisplayFormat("dd-MM-yyyy")
        fields_layout.addWidget(self.date_input, 1, 1)
        
        # Description field (optional)
        fields_layout.addWidget(QLabel("Description (Optional):"), 2, 0)
        self.description_input = QLineEdit()
        self.description_input.setPlaceholderText("e.g., Processing fee, Annual charge")
        fields_layout.addWidget(self.description_input, 2, 1)
        
        fields_group.setLayout(fields_layout)
        layout.addWidget(fields_group)
        
        # Buttons
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | 
                                     QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def get_charge_data(self):
        date = self.date_input.date()
        return {
            'amount': self.amount_input.value(),
            'date': datetime(date.year(), date.month(), date.day()),
            'description': self.description_input.text()
        }


class PrePaymentDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Add Pre-Payment")
        self.setModal(True)
        self.setMinimumWidth(500)
        
        layout = QVBoxLayout(self)
        
        # Type selection
        type_group = QGroupBox("Pre-Payment Type")
        type_layout = QVBoxLayout()
        
        self.type_combo = QComboBox()
        self.type_combo.addItems([
            "Single Date Payment",
            "Recurring Monthly Payment"
        ])
        self.type_combo.currentIndexChanged.connect(self.update_fields)
        type_layout.addWidget(self.type_combo)
        type_group.setLayout(type_layout)
        layout.addWidget(type_group)
        
        # Input fields group
        self.fields_group = QGroupBox("Payment Details")
        self.fields_layout = QGridLayout()
        self.fields_group.setLayout(self.fields_layout)
        layout.addWidget(self.fields_group)
        
        # Amount field (common for all types)
        self.fields_layout.addWidget(QLabel("Pre-Payment Amount:"), 0, 0)
        self.amount_input = QDoubleSpinBox()
        self.amount_input.setRange(0, 99999999)
        self.amount_input.setDecimals(2)
        self.amount_input.setValue(10000)
        self.amount_input.setPrefix("₹")
        self.fields_layout.addWidget(self.amount_input, 0, 1)
        
        # Single date fields
        self.single_date_label = QLabel("Payment Date:")
        self.single_date_input = QDateEdit()
        self.single_date_input.setCalendarPopup(True)
        self.single_date_input.setDate(QDate.currentDate())
        self.single_date_input.setDisplayFormat("dd-MM-yyyy")
        
        # Recurring fields
        self.recurring_day_label = QLabel("Day of Month:")
        self.recurring_day_input = QSpinBox()
        self.recurring_day_input.setRange(1, 31)
        self.recurring_day_input.setValue(15)
        
        self.recurring_start_label = QLabel("Start Date:")
        self.recurring_start_input = QDateEdit()
        self.recurring_start_input.setCalendarPopup(True)
        self.recurring_start_input.setDate(QDate.currentDate())
        self.recurring_start_input.setDisplayFormat("dd-MM-yyyy")
        
        self.recurring_end_label = QLabel("End Date (Optional):")
        self.recurring_end_input = QDateEdit()
        self.recurring_end_input.setCalendarPopup(True)
        self.recurring_end_input.setDate(QDate.currentDate().addYears(1))
        self.recurring_end_input.setDisplayFormat("dd-MM-yyyy")
        self.recurring_end_input.setSpecialValueText("No End Date")
        
        # Buttons
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | 
                                     QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
        
        self.update_fields()
        
    def update_fields(self):
        # Clear existing fields
        for i in reversed(range(1, self.fields_layout.rowCount())):
            for j in range(self.fields_layout.columnCount()):
                item = self.fields_layout.itemAtPosition(i, j)
                if item:
                    widget = item.widget()
                    if widget:
                        widget.setParent(None)
        
        payment_type = self.type_combo.currentText()
        
        if payment_type == "Single Date Payment":
            self.fields_layout.addWidget(self.single_date_label, 1, 0)
            self.fields_layout.addWidget(self.single_date_input, 1, 1)
            
        elif payment_type == "Recurring Monthly Payment":
            self.fields_layout.addWidget(self.recurring_day_label, 1, 0)
            self.fields_layout.addWidget(self.recurring_day_input, 1, 1)
            self.fields_layout.addWidget(self.recurring_start_label, 2, 0)
            self.fields_layout.addWidget(self.recurring_start_input, 2, 1)
            self.fields_layout.addWidget(self.recurring_end_label, 3, 0)
            self.fields_layout.addWidget(self.recurring_end_input, 3, 1)
    
    def get_prepayment_data(self):
        payment_type = self.type_combo.currentText()
        amount = self.amount_input.value()
        
        if payment_type == "Single Date Payment":
            date = self.single_date_input.date()
            return {
                'type': 'single',
                'amount': amount,
                'date': datetime(date.year(), date.month(), date.day())
            }
            
        elif payment_type == "Recurring Monthly Payment":
            start = self.recurring_start_input.date()
            end = self.recurring_end_input.date()
            return {
                'type': 'recurring',
                'amount': amount,
                'day': self.recurring_day_input.value(),
                'start_date': datetime(start.year(), start.month(), start.day()),
                'end_date': datetime(end.year(), end.month(), end.day()) if end else None
            }

class InterestRateRevisionDialog(QDialog):
    def __init__(self, existing_revisions=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Interest Rate Revisions")
        self.setModal(True)
        self.setMinimumWidth(450)
        
        layout = QVBoxLayout(self)
        
        # Instructions
        info_label = QLabel("Add interest rate revisions that will apply from specific dates:")
        info_label.setWordWrap(True)
        layout.addWidget(info_label)
        
        # Input fields group
        fields_group = QGroupBox("Revision Details")
        fields_layout = QGridLayout()
        
        # New APR field
        fields_layout.addWidget(QLabel("New APR (%):"), 0, 0)
        self.apr_input = QDoubleSpinBox()
        self.apr_input.setRange(0, 100)
        self.apr_input.setDecimals(2)
        self.apr_input.setValue(8.65)
        self.apr_input.setSuffix("%")
        fields_layout.addWidget(self.apr_input, 0, 1)
        
        # Effective date field
        fields_layout.addWidget(QLabel("Effective From:"), 1, 0)
        self.date_input = QDateEdit()
        self.date_input.setCalendarPopup(True)
        self.date_input.setDate(QDate.currentDate())
        self.date_input.setDisplayFormat("dd-MM-yyyy")
        fields_layout.addWidget(self.date_input, 1, 1)
        
        fields_group.setLayout(fields_layout)
        layout.addWidget(fields_group)
        
        # Add button
        add_btn = QPushButton("Add Revision")
        add_btn.clicked.connect(self.add_revision)
        layout.addWidget(add_btn)
        
        # List of revisions
        layout.addWidget(QLabel("Interest Rate Revisions:"))
        self.revision_list = QTextEdit()
        self.revision_list.setReadOnly(True)
        self.revision_list.setMaximumHeight(150)
        layout.addWidget(self.revision_list)
        
        # Initialize revisions
        self.revisions = existing_revisions if existing_revisions else []
        self.update_revision_display()
        
        # Remove button
        remove_btn = QPushButton("Remove Last Entry")
        remove_btn.clicked.connect(self.remove_last_revision)
        layout.addWidget(remove_btn)
        
        # Buttons
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | 
                                     QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def add_revision(self):
        apr = self.apr_input.value()
        date = self.date_input.date()
        
        revision = {
            'apr': apr,
            'date': datetime(date.year(), date.month(), date.day())
        }
        
        # Check if already exists for same date
        existing = False
        for i, rev in enumerate(self.revisions):
            if rev['date'] == revision['date']:
                self.revisions[i] = revision
                existing = True
                break
        
        if not existing:
            self.revisions.append(revision)
        
        self.update_revision_display()
    
    def remove_last_revision(self):
        if self.revisions:
            self.revisions.pop()
            self.update_revision_display()
    
    def update_revision_display(self):
        if not self.revisions:
            self.revision_list.setPlainText("No interest rate revisions added yet.")
            return
        
        # Sort by date
        sorted_revisions = sorted(self.revisions, key=lambda x: x['date'])
        
        text = ""
        for rev in sorted_revisions:
            text += f"{rev['date'].strftime('%d-%m-%Y')}: {rev['apr']}%\n"
        
        self.revision_list.setPlainText(text)
    
    def get_revisions(self):
        return sorted(self.revisions, key=lambda x: x['date'])

class LoanCalculatorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Loan Calculator Pro")
        self.setGeometry(100, 100, 1400, 850)
        
        # Initialize prepayments, bank charges, manual EMIs, and EMI exclusions
        self.prepayments = []
        self.bank_charges = []
        self.manual_emis = []
        self.emi_exclusions = []
        self.interest_rate_revisions = []
        
        # Set modern stylesheet
        # Set modern stylesheet
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f7fa;
            }
            QGroupBox {
                font-size: 16px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 10px;
                margin-top: 15px;
                padding-top: 20px;
                background-color: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top center;
                padding: 5px 15px;
                background-color: #3498db;
                color: white;
                border-radius: 5px;
            }
            QLabel {
                color: #34495e;
                font-size: 13px;
                font-weight: 500;
            }
            QLineEdit {
                padding: 8px 12px;
                border: 2px solid #bdc3c7;
                border-radius: 6px;
                background-color: #ffffff;
                font-size: 13px;
                color: #2c3e50;
            }
            QLineEdit:focus {
                border: 2px solid #3498db;
                background-color: #f8fbff;
            }
            QComboBox {
                padding: 8px 12px;
                border: 2px solid #bdc3c7;
                border-radius: 6px;
                background-color: #ffffff;
                font-size: 13px;
                color: #2c3e50;
            }
            QComboBox:focus {
                border: 2px solid #3498db;
                background-color: #f8fbff;
            }
            QComboBox::drop-down {
                border: none;
                width: 30px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 6px solid #2c3e50;
                margin-right: 5px;
            }
            QDateEdit {
                padding: 8px 12px;
                border: 2px solid #bdc3c7;
                border-radius: 6px;
                background-color: #ffffff;
                font-size: 13px;
                color: #2c3e50;
            }
            QDateEdit:focus {
                border: 2px solid #3498db;
                background-color: #f8fbff;
            }
            QSpinBox, QDoubleSpinBox {
                padding: 8px 12px;
                border: 2px solid #bdc3c7;
                border-radius: 6px;
                background-color: #ffffff;
                font-size: 13px;
                color: #2c3e50;
            }
            QSpinBox:focus, QDoubleSpinBox:focus {
                border: 2px solid #3498db;
                background-color: #f8fbff;
            }
            QTextEdit {
                padding: 8px 12px;
                border: 2px solid #bdc3c7;
                border-radius: 6px;
                background-color: #ffffff;
                font-size: 13px;
                color: #2c3e50;
            }
            QTextEdit:focus {
                border: 2px solid #3498db;
                background-color: #f8fbff;
            }
            QDialog {
                background-color: #f5f7fa;
            }
            QPushButton {
                padding: 12px 30px;
                font-size: 14px;
                font-weight: bold;
                border: none;
                border-radius: 8px;
                color: white;
                background-color: #3498db;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #21618c;
            }
            QPushButton#calculateBtn {
                background-color: #27ae60;
            }
            QPushButton#calculateBtn:hover {
                background-color: #229954;
            }
            QPushButton#calculateBtn:pressed {
                background-color: #1e8449;
            }
            QPushButton#exportBtn {
                background-color: #9b59b6;
            }
            QPushButton#exportBtn:hover {
                background-color: #8e44ad;
            }
            QPushButton#exportBtn:pressed {
                background-color: #7d3c98;
            }
            QPushButton#clearBtn {
                background-color: #e74c3c;
            }
            QPushButton#clearBtn:hover {
                background-color: #c0392b;
            }
            QPushButton#clearBtn:pressed {
                background-color: #a93226;
            }
            QDialogButtonBox QPushButton {
                min-width: 80px;
                padding: 10px 20px;
            }
            QTableWidget {
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                background-color: #ffffff;
                gridline-color: #ecf0f1;
                font-size: 12px;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QHeaderView::section {
                background-color: #3498db;
                color: white;
                padding: 8px;
                border: none;
                font-weight: bold;
                font-size: 11px;
            }
            QTabWidget::pane {
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #ecf0f1;
                color: #2c3e50;
                padding: 10px 20px;
                margin-right: 2px;
                border-top-left-radius: 6px;
                border-top-right-radius: 6px;
                font-weight: bold;
            }
            QTabBar::tab:selected {
                background-color: #3498db;
                color: white;
            }
        """)
        
        # Create central widget and main layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        # Title
        title_label = QLabel("💰 Loan Calculator")
        title_label.setStyleSheet("""
            font-size: 28px;
            font-weight: bold;
            color: #2c3e50;
            padding: 10px;
        """)
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(title_label)
        
        # Create input fields section
        # Create input fields section with collapse/expand
        input_container = QWidget()
        input_container_layout = QVBoxLayout(input_container)
        input_container_layout.setContentsMargins(0, 0, 0, 0)
        input_container_layout.setSpacing(0)
        
        # Header with collapse button
        input_header = QWidget()
        input_header.setStyleSheet("""
            QWidget {
                background-color: white;
                border: 2px solid #3498db;
                border-radius: 10px;
            }
        """)
        input_header_layout = QHBoxLayout(input_header)
        input_header_layout.setContentsMargins(20, 10, 20, 10)
        
        input_title = QLabel("📝 Input Fields")
        input_title.setStyleSheet("""
            font-size: 16px;
            font-weight: bold;
            color: #2c3e50;
            background-color: transparent;
            border: none;
        """)
        
        self.collapse_btn = QPushButton("▼ Collapse")
        self.collapse_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px 15px;
                font-size: 12px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        self.collapse_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.collapse_btn.clicked.connect(self.toggle_input_fields)
        
        input_header_layout.addWidget(input_title)
        input_header_layout.addStretch()
        input_header_layout.addWidget(self.collapse_btn)
        
        # Input fields group
        self.input_group = QGroupBox()
        self.input_group.setStyleSheet("""
            QGroupBox {
                border: 2px solid #3498db;
                border-top: none;
                border-top-left-radius: 0px;
                border-top-right-radius: 0px;
                border-bottom-left-radius: 10px;
                border-bottom-right-radius: 10px;
                margin-top: 0px;
                padding-top: 20px;
                background-color: white;
            }
        """)
        input_layout = QGridLayout()
        input_layout.setSpacing(15)
        input_layout.setContentsMargins(20, 25, 20, 20)
        
        # Row 0: Loan Amount and APR
        input_layout.addWidget(self.create_label("Loan Amount:"), 0, 0)
        self.loan_amount = QLineEdit()
        self.loan_amount.setText("5000000.00")
        self.loan_amount.setPlaceholderText("Enter loan amount")
        input_layout.addWidget(self.loan_amount, 0, 1)
        
        input_layout.addWidget(self.create_label("APR (%):"), 0, 3)
        self.apr = QLineEdit()
        self.apr.setText("8.65")
        self.apr.setPlaceholderText("Annual percentage rate")
        input_layout.addWidget(self.apr, 0, 4)
        
        input_layout.addWidget(self.create_label("Year Base:"), 0, 6)
        self.year_base = QLineEdit()
        self.year_base.setText("365")
        self.year_base.setPlaceholderText("Days in year")
        input_layout.addWidget(self.year_base, 0, 7)
        
        # Bank Charges section
        input_layout.addWidget(self.create_label("Bank Charges:"), 0, 9)
        bank_charge_buttons = QWidget()
        bank_charge_layout = QHBoxLayout(bank_charge_buttons)
        bank_charge_layout.setContentsMargins(0, 0, 0, 0)
        bank_charge_layout.setSpacing(5)
        
        self.add_bank_charge_btn = QPushButton("+ Add")
        self.add_bank_charge_btn.setStyleSheet("""
            QPushButton {
                background-color: #e67e22;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #d35400;
            }
        """)
        self.add_bank_charge_btn.clicked.connect(self.add_bank_charge)
        
        self.view_bank_charges_btn = QPushButton("View (0)")
        self.view_bank_charges_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        self.view_bank_charges_btn.clicked.connect(self.view_bank_charges)
        
        self.clear_bank_charges_btn = QPushButton("Clear")
        self.clear_bank_charges_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        self.clear_bank_charges_btn.clicked.connect(self.clear_bank_charges)
        
        bank_charge_layout.addWidget(self.add_bank_charge_btn)
        bank_charge_layout.addWidget(self.view_bank_charges_btn)
        bank_charge_layout.addWidget(self.clear_bank_charges_btn)
        bank_charge_layout.addStretch()
        
        input_layout.addWidget(bank_charge_buttons, 0, 10)
        
        # Row 1: Loan Start Date, EMI, and EMI Date
        input_layout.addWidget(self.create_label("Loan Start Date:"), 1, 0)
        
        # Create date field with calendar button
        date_widget = QWidget()
        date_widget_layout = QHBoxLayout(date_widget)
        date_widget_layout.setContentsMargins(0, 0, 0, 0)
        date_widget_layout.setSpacing(0)
        
        self.loan_start_dt = QDateEdit()
        self.loan_start_dt.setCalendarPopup(False)
        self.loan_start_dt.setDate(QDate(2024, 5, 2))
        self.loan_start_dt.setDisplayFormat("dd-MM-yyyy")
        self.loan_start_dt.setButtonSymbols(QDateEdit.ButtonSymbols.NoButtons)
        self.loan_start_dt.setReadOnly(True)
        self.loan_start_dt.setStyleSheet("""
            QDateEdit {
                padding: 8px 12px;
                border: 2px solid #bdc3c7;
                border-top-left-radius: 6px;
                border-bottom-left-radius: 6px;
                border-right: none;
                background-color: #ffffff;
                font-size: 13px;
                color: #2c3e50;
                min-width: 140px;
            }
            QDateEdit:focus {
                border: 2px solid #3498db;
                border-right: none;
                background-color: #f8fbff;
            }
        """)
        
        # Create custom calendar button
        calendar_btn = QPushButton("📅")
        calendar_btn.setFixedSize(45, 38)
        calendar_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        calendar_btn.setStyleSheet("""
            QPushButton {
                background-color: #ecf0f1;
                border: 2px solid #bdc3c7;
                border-left: 1px solid #bdc3c7;
                border-top-right-radius: 6px;
                border-bottom-right-radius: 6px;
                font-size: 20px;
                padding: 0px;
            }
            QPushButton:hover {
                background-color: #d5dbdb;
            }
            QPushButton:pressed {
                background-color: #bdc3c7;
            }
        """)
        
        def open_calendar():
            from PyQt6.QtWidgets import QCalendarWidget
            
            if not hasattr(self, 'calendar_popup') or self.calendar_popup is None:
                self.calendar_popup = QCalendarWidget()
                self.calendar_popup.setWindowFlags(Qt.WindowType.Popup)
                self.calendar_popup.clicked.connect(lambda date: (
                    self.loan_start_dt.setDate(date),
                    self.calendar_popup.hide()
                ))
            
            self.calendar_popup.setSelectedDate(self.loan_start_dt.date())
            pos = self.loan_start_dt.mapToGlobal(self.loan_start_dt.rect().bottomLeft())
            self.calendar_popup.move(pos)
            self.calendar_popup.show()
            self.calendar_popup.setFocus()
        
        calendar_btn.clicked.connect(open_calendar)
        
        date_widget_layout.addWidget(self.loan_start_dt)
        date_widget_layout.addWidget(calendar_btn)
        date_widget_layout.addStretch()
        
        input_layout.addWidget(date_widget, 1, 1)
        
        input_layout.addWidget(self.create_label("EMI Amount:"), 1, 3)
        self.emi = QLineEdit()
        self.emi.setText("40800.00")
        self.emi.setPlaceholderText("Monthly EMI")
        input_layout.addWidget(self.emi, 1, 4)
        
        input_layout.addWidget(self.create_label("EMI Date:"), 1, 6)
        self.emi_date = QLineEdit()
        self.emi_date.setText("5")
        self.emi_date.setPlaceholderText("Day of month")
        input_layout.addWidget(self.emi_date, 1, 7)
        
        # Exclude EMI Months button
        exclude_emi_widget = QWidget()
        exclude_emi_layout = QHBoxLayout(exclude_emi_widget)
        exclude_emi_layout.setContentsMargins(0, 0, 0, 0)
        exclude_emi_layout.setSpacing(5)
        
        self.exclude_emi_btn = QPushButton("Ex (0)")
        self.exclude_emi_btn.setStyleSheet("""
            QPushButton {
                background-color: #8e44ad;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #7d3c98;
            }
        """)
        self.exclude_emi_btn.clicked.connect(self.manage_emi_exclusions)
        exclude_emi_layout.addWidget(self.exclude_emi_btn)
        exclude_emi_layout.addStretch()
        
        input_layout.addWidget(exclude_emi_widget, 1, 8)
        
        # Manual EMI section
        input_layout.addWidget(self.create_label("Manual EMI:"), 1, 9)
        manual_emi_buttons = QWidget()
        manual_emi_layout = QHBoxLayout(manual_emi_buttons)
        manual_emi_layout.setContentsMargins(0, 0, 0, 0)
        manual_emi_layout.setSpacing(5)
        
        self.add_manual_emi_btn = QPushButton("+ Add")
        self.add_manual_emi_btn.setStyleSheet("""
            QPushButton {
                background-color: #16a085;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #138d75;
            }
        """)
        self.add_manual_emi_btn.clicked.connect(self.add_manual_emi)
        
        self.view_manual_emis_btn = QPushButton("View (0)")
        self.view_manual_emis_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        self.view_manual_emis_btn.clicked.connect(self.view_manual_emis)
        
        self.clear_manual_emis_btn = QPushButton("Clear")
        self.clear_manual_emis_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        self.clear_manual_emis_btn.clicked.connect(self.clear_manual_emis)
        
        manual_emi_layout.addWidget(self.add_manual_emi_btn)
        manual_emi_layout.addWidget(self.view_manual_emis_btn)
        manual_emi_layout.addWidget(self.clear_manual_emis_btn)
        manual_emi_layout.addStretch()
        
        input_layout.addWidget(manual_emi_buttons, 1, 10)
        
        # Interest Rate Revision section
        input_layout.addWidget(self.create_label("Interest Rate Revisions:"), 2, 9)
        rate_revision_buttons = QWidget()
        rate_revision_layout = QHBoxLayout(rate_revision_buttons)
        rate_revision_layout.setContentsMargins(0, 0, 0, 0)
        rate_revision_layout.setSpacing(5)
        
        self.add_rate_revision_btn = QPushButton("+ Add")
        self.add_rate_revision_btn.setStyleSheet("""
            QPushButton {
                background-color: #9b59b6;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #8e44ad;
            }
        """)
        self.add_rate_revision_btn.clicked.connect(self.add_interest_rate_revision)
        
        self.view_rate_revisions_btn = QPushButton("View (0)")
        self.view_rate_revisions_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        self.view_rate_revisions_btn.clicked.connect(self.view_interest_rate_revisions)
        
        self.clear_rate_revisions_btn = QPushButton("Clear")
        self.clear_rate_revisions_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        self.clear_rate_revisions_btn.clicked.connect(self.clear_interest_rate_revisions)
        
        rate_revision_layout.addWidget(self.add_rate_revision_btn)
        rate_revision_layout.addWidget(self.view_rate_revisions_btn)
        rate_revision_layout.addWidget(self.clear_rate_revisions_btn)
        rate_revision_layout.addStretch()
        
        input_layout.addWidget(rate_revision_buttons, 2, 10)
        
        # Row 2: Interest Charged Date and Loan Tenure
        input_layout.addWidget(self.create_label("Interest Charged Date:"), 2, 0)
        
        self.interest_charged_date = QComboBox()
        # Add values 1-31 and EOM
        for i in range(1, 32):
            self.interest_charged_date.addItem(str(i))
        self.interest_charged_date.addItem("EOM")
        self.interest_charged_date.setCurrentText("5")
        input_layout.addWidget(self.interest_charged_date, 2, 1)
        
        input_layout.addWidget(self.create_label("Loan Tenure (Months):"), 2, 3)
        self.loan_tenure = QLineEdit()
        self.loan_tenure.setText("300")
        self.loan_tenure.setPlaceholderText("Total months")
        input_layout.addWidget(self.loan_tenure, 2, 4)
        
        # Pre-payment section
        input_layout.addWidget(self.create_label("Pre-Payments:"), 2, 6)
        prepayment_buttons = QWidget()
        prepayment_layout = QHBoxLayout(prepayment_buttons)
        prepayment_layout.setContentsMargins(0, 0, 0, 0)
        prepayment_layout.setSpacing(5)
        
        self.add_prepayment_btn = QPushButton("+ Add")
        self.add_prepayment_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        self.add_prepayment_btn.clicked.connect(self.add_prepayment)
        
        self.view_prepayments_btn = QPushButton("View (0)")
        self.view_prepayments_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        self.view_prepayments_btn.clicked.connect(self.view_prepayments)
        
        self.clear_prepayments_btn = QPushButton("Clear")
        self.clear_prepayments_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        self.clear_prepayments_btn.clicked.connect(self.clear_prepayments)
        
        prepayment_layout.addWidget(self.add_prepayment_btn)
        prepayment_layout.addWidget(self.view_prepayments_btn)
        prepayment_layout.addWidget(self.clear_prepayments_btn)
        prepayment_layout.addStretch()
        
        input_layout.addWidget(prepayment_buttons, 2, 7)
        
        # Set column stretches
        input_layout.setColumnStretch(2, 1)
        input_layout.setColumnStretch(5, 1)
        input_layout.setColumnMinimumWidth(1, 150)
        input_layout.setColumnMinimumWidth(4, 150)
        input_layout.setColumnMinimumWidth(7, 100)
        
        self.input_group.setLayout(input_layout)
        
        input_container_layout.addWidget(input_header)
        input_container_layout.addWidget(self.input_group)
        
        main_layout.addWidget(input_container)
        
        # Create output section with tabs
        output_group = QGroupBox("📊 Output Results")
        output_layout = QVBoxLayout()
        output_layout.setContentsMargins(20, 25, 20, 20)
        
        # Create tab widget
        self.tab_widget = QTabWidget()
        
        # Summary tab
        # Summary tab
        self.summary_text = QTextEdit()
        self.summary_text.setReadOnly(True)
        # Remove or increase the maximum height
        # self.summary_text.setMaximumHeight(150)  # Comment out or remove this line
        self.tab_widget.addTab(self.summary_text, "📋 Summary")
        
        # Amortization schedule table
        self.schedule_table = QTableWidget()
        self.schedule_table.setColumnCount(15)
        self.schedule_table.setHorizontalHeaderLabels([
            "Date", "Beginning\nBalance", "Misc Charges\nBy Bank", 
            "Interest Rate\n(Annual)", "Interest Rate\n(Daily)", 
            "Interest Amount\nfor Each Day", "Cumulative\nInterest\nAccrued",
            "Interest\nDebited By\nBank", "EMI", "Pre-Payment",
            "Interest Paid", "Principal\nPaid", "Remaining\nBalance",
            "Balance +\nInterest Due", "Total Interest\nPaid"
        ])
        
        # Set table properties
        self.schedule_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.schedule_table.setAlternatingRowColors(True)
        self.schedule_table.verticalHeader().setVisible(False)
        self.schedule_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.schedule_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.schedule_table.cellDoubleClicked.connect(self.edit_prepayment_cell)
        
        self.tab_widget.addTab(self.schedule_table, "📅 Amortization Schedule")
        
        output_layout.addWidget(self.tab_widget)
        output_group.setLayout(output_layout)
        main_layout.addWidget(output_group)
        
        # Create button section
# Create button section
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)
        
        self.calculate_btn = QPushButton("🔢 Calculate")
        self.calculate_btn.setObjectName("calculateBtn")
        self.calculate_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.calculate_btn.clicked.connect(self.calculate)
        
        self.export_btn = QPushButton("📊 Export to Excel")
        self.export_btn.setObjectName("exportBtn")
        self.export_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.export_btn.clicked.connect(self.export_to_excel)
        
        self.clear_btn = QPushButton("🗑️ Clear All")
        self.clear_btn.setObjectName("clearBtn")
        self.clear_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.clear_btn.clicked.connect(self.clear_fields)
        
        # Credit label
        credit_label = QLabel("Developed by Yogesh Khurana")
        credit_label.setStyleSheet("""
            font-size: 14px;
            color: #7f8c8d;
            font-style: italic;
            padding-right: 10px;
        """)
        credit_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        
        button_layout.addStretch(3)  # Smaller stretch on left
        button_layout.addWidget(self.calculate_btn)
        button_layout.addWidget(self.export_btn)
        button_layout.addWidget(self.clear_btn)
        button_layout.addStretch(2)  # Larger stretch in middle
        button_layout.addWidget(credit_label)
        
        main_layout.addLayout(button_layout)
        self.load_settings()
    
    def toggle_input_fields(self):
        """Toggle collapse/expand of input fields"""
        if self.input_group.isVisible():
            # Collapse
            self.input_group.hide()
            self.collapse_btn.setText("▶ Expand")
        else:
            # Expand
            self.input_group.show()
            self.collapse_btn.setText("▼ Collapse")    
    
    def create_label(self, text):
        """Create a styled label"""
        return QLabel(text)
    
    def manage_emi_exclusions(self):
        """Open dialog to manage EMI exclusions"""
        dialog = ExcludeMonthsDialog(self.emi_exclusions, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.emi_exclusions = dialog.get_exclusions()
            self.exclude_emi_btn.setText(f"Ex ({len(self.emi_exclusions)})")
    
    def is_emi_excluded(self, date):
        """Check if EMI is excluded for this date's month/year"""
        for exclusion in self.emi_exclusions:
            if date.month == exclusion['month'] and date.year == exclusion['year']:
                return True
        return False
    
    def add_manual_emi(self):
        """Open dialog to add manual EMI"""
        dialog = ManualEMIDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            emi_data = dialog.get_emi_data()
            self.manual_emis.append(emi_data)
            self.view_manual_emis_btn.setText(f"View ({len(self.manual_emis)})")
    
    def view_manual_emis(self):
        """View all manual EMIs with delete option"""
        if not self.manual_emis:
            dialog = QDialog(self)
            dialog.setWindowTitle("Manual EMIs")
            layout = QVBoxLayout(dialog)
            layout.addWidget(QLabel("No manual EMIs added yet."))
            button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
            button_box.accepted.connect(dialog.accept)
            layout.addWidget(button_box)
            dialog.exec()
            return
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Manual EMIs List")
        dialog.setMinimumWidth(600)
        dialog.setMinimumHeight(400)
        layout = QVBoxLayout(dialog)
        
        # Create table
        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["Date", "Amount", "Note", "Action"])
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        
        def refresh_table():
            """Refresh the table content"""
            table.setRowCount(len(self.manual_emis))
            for i, emi in enumerate(self.manual_emis):
                # Date
                date_item = QTableWidgetItem(emi['date'].strftime('%d-%m-%Y'))
                date_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                table.setItem(i, 0, date_item)
                
                # Amount
                amount_item = QTableWidgetItem(f"₹{emi['amount']:,.2f}")
                amount_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                table.setItem(i, 1, amount_item)
                
                # Note
                note_item = QTableWidgetItem(emi['note'] if emi['note'] else "-")
                table.setItem(i, 2, note_item)
                
                # Delete button
                delete_btn = QPushButton("🗑️ Delete")
                delete_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #e74c3c;
                        color: white;
                        border: none;
                        border-radius: 9px;
                        padding: 5px 14px;
                        font-size: 11px;
                        min-width: 50px;
                    }
                    QPushButton:hover {
                        background-color: #c0392b;
                    }
                """)
                delete_btn.clicked.connect(lambda checked, idx=i: self.delete_manual_emi(idx, dialog, refresh_table))
                table.setCellWidget(i, 3, delete_btn)
        
        refresh_table()
        layout.addWidget(table)
        
        # OK button
        button_layout = QHBoxLayout()
        ok_btn = QPushButton("OK")
        ok_btn.clicked.connect(dialog.accept)
        button_layout.addStretch()
        button_layout.addWidget(ok_btn)
        layout.addLayout(button_layout)
        
        dialog.exec()
    
    def delete_manual_emi(self, index, dialog, refresh_callback):
        """Delete a manual EMI"""
        from PyQt6.QtWidgets import QMessageBox
        reply = QMessageBox.question(
            dialog,
            "Confirm Delete",
            f"Are you sure you want to delete this manual EMI?\n\n"
            f"Date: {self.manual_emis[index]['date'].strftime('%d-%m-%Y')}\n"
            f"Amount: ₹{self.manual_emis[index]['amount']:,.2f}",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.manual_emis.pop(index)
            self.view_manual_emis_btn.setText(f"View ({len(self.manual_emis)})")
            # Check if list is empty
            if not self.manual_emis:
                dialog.accept()  # Close dialog if no items left
            else:
                refresh_callback()  # Refresh the table    
    def clear_manual_emis(self):
        """Clear all manual EMIs"""
        self.manual_emis = []
        self.view_manual_emis_btn.setText("View (0)")
    
    def get_manual_emi_for_date(self, date):
        """Get total manual EMI amount for a specific date"""
        total = 0
        for emi in self.manual_emis:
            if emi['date'].date() == date.date():
                total += emi['amount']
        return total
    
    def add_bank_charge(self):
        """Open dialog to add bank charge"""
        dialog = BankChargeDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            charge_data = dialog.get_charge_data()
            self.bank_charges.append(charge_data)
            self.view_bank_charges_btn.setText(f"View ({len(self.bank_charges)})")
    
    def view_bank_charges(self):
        """View all bank charges with delete option"""
        if not self.bank_charges:
            dialog = QDialog(self)
            dialog.setWindowTitle("Bank Charges")
            layout = QVBoxLayout(dialog)
            layout.addWidget(QLabel("No bank charges added yet."))
            button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
            button_box.accepted.connect(dialog.accept)
            layout.addWidget(button_box)
            dialog.exec()
            return
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Bank Charges List")
        dialog.setMinimumWidth(600)
        dialog.setMinimumHeight(400)
        layout = QVBoxLayout(dialog)
        
        # Create table
        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["Date", "Amount", "Description", "Action"])
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        
        def refresh_table():
            """Refresh the table content"""
            table.setRowCount(len(self.bank_charges))
            for i, charge in enumerate(self.bank_charges):
                # Date
                date_item = QTableWidgetItem(charge['date'].strftime('%d-%m-%Y'))
                date_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                table.setItem(i, 0, date_item)
                
                # Amount
                amount_item = QTableWidgetItem(f"₹{charge['amount']:,.2f}")
                amount_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                table.setItem(i, 1, amount_item)
                
                # Description
                desc_item = QTableWidgetItem(charge['description'] if charge['description'] else "-")
                table.setItem(i, 2, desc_item)
                
                # Delete button
                delete_btn = QPushButton("🗑️ Delete")
                delete_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #e74c3c;
                        color: white;
                        border: none;
                        border-radius: 9px;
                        padding: 5px 14px;
                        font-size: 11px;
                        min-width: 50px;
                    }
                    QPushButton:hover {
                        background-color: #c0392b;
                    }
                """)
                delete_btn.clicked.connect(lambda checked, idx=i: self.delete_bank_charge(idx, dialog, refresh_table))
                table.setCellWidget(i, 3, delete_btn)
        
        refresh_table()
        layout.addWidget(table)
        
        # OK button
        button_layout = QHBoxLayout()
        ok_btn = QPushButton("OK")
        ok_btn.clicked.connect(dialog.accept)
        button_layout.addStretch()
        button_layout.addWidget(ok_btn)
        layout.addLayout(button_layout)
        
        dialog.exec()
    
    def delete_bank_charge(self, index, dialog, refresh_callback):
        """Delete a bank charge"""
        from PyQt6.QtWidgets import QMessageBox
        reply = QMessageBox.question(
            dialog,
            "Confirm Delete",
            f"Are you sure you want to delete this bank charge?\n\n"
            f"Date: {self.bank_charges[index]['date'].strftime('%d-%m-%Y')}\n"
            f"Amount: ₹{self.bank_charges[index]['amount']:,.2f}",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.bank_charges.pop(index)
            self.view_bank_charges_btn.setText(f"View ({len(self.bank_charges)})")
            # Check if list is empty
            if not self.bank_charges:
                dialog.accept()  # Close dialog if no items left
            else:
                refresh_callback()  # Refresh the table    
    def clear_bank_charges(self):
        """Clear all bank charges"""
        self.bank_charges = []
        self.view_bank_charges_btn.setText("View (0)")
    
    def get_bank_charge_for_date(self, date):
        """Get total bank charge amount for a specific date"""
        total = 0
        for charge in self.bank_charges:
            if charge['date'].date() == date.date():
                total += charge['amount']
        return total
    
    def add_prepayment(self):
        """Open dialog to add prepayment"""
        dialog = PrePaymentDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            prepayment_data = dialog.get_prepayment_data()
            self.prepayments.append(prepayment_data)
            self.view_prepayments_btn.setText(f"View ({len(self.prepayments)})")
    
    def view_prepayments(self):
        """View all prepayments with delete option"""
        if not self.prepayments:
            dialog = QDialog(self)
            dialog.setWindowTitle("Pre-Payments")
            layout = QVBoxLayout(dialog)
            layout.addWidget(QLabel("No pre-payments added yet."))
            button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
            button_box.accepted.connect(dialog.accept)
            layout.addWidget(button_box)
            dialog.exec()
            return
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Pre-Payments List")
        dialog.setMinimumWidth(700)
        dialog.setMinimumHeight(400)
        layout = QVBoxLayout(dialog)
        
        # Create table
        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["Type", "Amount", "Details", "Action"])
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        
        def refresh_table():
            """Refresh the table content"""
            table.setRowCount(len(self.prepayments))
            for i, pp in enumerate(self.prepayments):
                # Type
                type_item = QTableWidgetItem("Single" if pp['type'] == 'single' else "Recurring")
                type_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                table.setItem(i, 0, type_item)
                
                # Amount
                amount_item = QTableWidgetItem(f"₹{pp['amount']:,.2f}")
                amount_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                table.setItem(i, 1, amount_item)
                
                # Details
                if pp['type'] == 'single':
                    details = f"Date: {pp['date'].strftime('%d-%m-%Y')}"
                else:
                    end_str = pp['end_date'].strftime('%d-%m-%Y') if pp['end_date'] else "No End"
                    details = f"Day {pp['day']} | From: {pp['start_date'].strftime('%d-%m-%Y')} | To: {end_str}"
                
                details_item = QTableWidgetItem(details)
                table.setItem(i, 2, details_item)
                
                # Delete button
                delete_btn = QPushButton("🗑️ Delete")
                delete_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #e74c3c;
                        color: white;
                        border: none;
                        border-radius: 9px;
                        padding: 5px 14px;
                        font-size: 11px;
                        min-width: 50px;
                    }
                    QPushButton:hover {
                        background-color: #c0392b;
                    }
                """)
                delete_btn.clicked.connect(lambda checked, idx=i: self.delete_prepayment(idx, dialog, refresh_table))
                table.setCellWidget(i, 3, delete_btn)
        
        refresh_table()
        layout.addWidget(table)
        
        # OK button
        button_layout = QHBoxLayout()
        ok_btn = QPushButton("OK")
        ok_btn.clicked.connect(dialog.accept)
        button_layout.addStretch()
        button_layout.addWidget(ok_btn)
        layout.addLayout(button_layout)
        
        dialog.exec()
    
    def delete_prepayment(self, index, dialog, refresh_callback):
        """Delete a prepayment"""
        from PyQt6.QtWidgets import QMessageBox
        pp = self.prepayments[index]
        
        if pp['type'] == 'single':
            details = f"Date: {pp['date'].strftime('%d-%m-%Y')}\nAmount: ₹{pp['amount']:,.2f}"
        else:
            details = f"Recurring on day {pp['day']}\nAmount: ₹{pp['amount']:,.2f}"
        
        reply = QMessageBox.question(
            dialog,
            "Confirm Delete",
            f"Are you sure you want to delete this pre-payment?\n\n{details}",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.prepayments.pop(index)
            self.view_prepayments_btn.setText(f"View ({len(self.prepayments)})")
            # Check if list is empty
            if not self.prepayments:
                dialog.accept()  # Close dialog if no items left
            else:
                refresh_callback()  # Refresh the table    
    def clear_prepayments(self):
        """Clear all prepayments"""
        self.prepayments = []
        self.view_prepayments_btn.setText("View (0)")
    
    def edit_prepayment_cell(self, row, col):
        """Allow editing prepayment in table cell"""
        if col != 9:  # Only allow editing Pre-Payment column
            return
        
        date_item = self.schedule_table.item(row, 0)
        if not date_item:
            return
        
        date_str = date_item.text()
        current_value_item = self.schedule_table.item(row, 9)
        current_value = 0
        if current_value_item and current_value_item.text():
            try:
                current_value = float(current_value_item.text().replace('₹', '').replace(',', ''))
            except:
                current_value = 0
        
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Edit Pre-Payment for {date_str}")
        layout = QVBoxLayout(dialog)
        
        layout.addWidget(QLabel(f"Enter pre-payment amount for {date_str}:"))
        
        amount_input = QDoubleSpinBox()
        amount_input.setRange(0, 99999999)
        amount_input.setDecimals(2)
        amount_input.setValue(current_value)
        amount_input.setPrefix("₹")
        layout.addWidget(amount_input)
        
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | 
                                     QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            amount = amount_input.value()
            
            # Parse the date
            date_obj = datetime.strptime(date_str, "%d-%m-%Y")
            
            # Add or update single prepayment for this date
            found = False
            for pp in self.prepayments:
                if pp['type'] == 'single' and pp['date'] == date_obj:
                    pp['amount'] = amount
                    found = True
                    break
            
            if not found and amount > 0:
                self.prepayments.append({
                    'type': 'single',
                    'amount': amount,
                    'date': date_obj
                })
            
            # Recalculate
            self.calculate()
    
    def get_prepayment_for_date(self, date):
        """Get total prepayment amount for a specific date"""
        total = 0
        
        for pp in self.prepayments:
            if pp['type'] == 'single':
                if pp['date'].date() == date.date():
                    total += pp['amount']
                    
            elif pp['type'] == 'recurring':
                if pp['start_date'].date() <= date.date():
                    if pp['end_date'] is None or date.date() <= pp['end_date'].date():
                        if date.day == pp['day']:
                            total += pp['amount']
        
        return total
        
    def add_interest_rate_revision(self):
        """Open dialog to add interest rate revision"""
        dialog = InterestRateRevisionDialog(self.interest_rate_revisions, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.interest_rate_revisions = dialog.get_revisions()
            self.view_rate_revisions_btn.setText(f"View ({len(self.interest_rate_revisions)})")
    
    def view_interest_rate_revisions(self):
        """View all interest rate revisions with delete option"""
        if not self.interest_rate_revisions:
            dialog = QDialog(self)
            dialog.setWindowTitle("Interest Rate Revisions")
            layout = QVBoxLayout(dialog)
            layout.addWidget(QLabel("No interest rate revisions added yet."))
            button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
            button_box.accepted.connect(dialog.accept)
            layout.addWidget(button_box)
            dialog.exec()
            return
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Interest Rate Revisions List")
        dialog.setMinimumWidth(600)
        dialog.setMinimumHeight(400)
        layout = QVBoxLayout(dialog)
        
        # Create table
        table = QTableWidget()
        table.setColumnCount(3)
        table.setHorizontalHeaderLabels(["Effective From", "New APR", "Action"])
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        
        def refresh_table():
            """Refresh the table content"""
            table.setRowCount(len(self.interest_rate_revisions))
            for i, rev in enumerate(self.interest_rate_revisions):
                # Date
                date_item = QTableWidgetItem(rev['date'].strftime('%d-%m-%Y'))
                date_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                table.setItem(i, 0, date_item)
                
                # APR
                apr_item = QTableWidgetItem(f"{rev['apr']}%")
                apr_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                table.setItem(i, 1, apr_item)
                
                # Delete button
                delete_btn = QPushButton("🗑️ Delete")
                delete_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #e74c3c;
                        color: white;
                        border: none;
                        border-radius: 9px;
                        padding: 5px 14px;
                        font-size: 11px;
                        min-width: 50px;
                    }
                    QPushButton:hover {
                        background-color: #c0392b;
                    }
                """)
                delete_btn.clicked.connect(lambda checked, idx=i: self.delete_interest_rate_revision(idx, dialog, refresh_table))
                table.setCellWidget(i, 2, delete_btn)
        
        refresh_table()
        layout.addWidget(table)
        
        # OK button
        button_layout = QHBoxLayout()
        ok_btn = QPushButton("OK")
        ok_btn.clicked.connect(dialog.accept)
        button_layout.addStretch()
        button_layout.addWidget(ok_btn)
        layout.addLayout(button_layout)
        
        dialog.exec()
    
    def delete_interest_rate_revision(self, index, dialog, refresh_callback):
        """Delete an interest rate revision"""
        from PyQt6.QtWidgets import QMessageBox
        reply = QMessageBox.question(
            dialog,
            "Confirm Delete",
            f"Are you sure you want to delete this interest rate revision?\n\n"
            f"Effective From: {self.interest_rate_revisions[index]['date'].strftime('%d-%m-%Y')}\n"
            f"APR: {self.interest_rate_revisions[index]['apr']}%",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.interest_rate_revisions.pop(index)
            self.view_rate_revisions_btn.setText(f"View ({len(self.interest_rate_revisions)})")
            # Check if list is empty
            if not self.interest_rate_revisions:
                dialog.accept()  # Close dialog if no items left
            else:
                refresh_callback()  # Refresh the table    
    def clear_interest_rate_revisions(self):
        """Clear all interest rate revisions"""
        self.interest_rate_revisions = []
        self.view_rate_revisions_btn.setText("View (0)")
    
    def get_apr_for_date(self, date):
        """Get applicable APR for a specific date based on revisions"""
        base_apr = float(self.apr.text())
        
        # If no revisions, return base APR
        if not self.interest_rate_revisions:
            return base_apr
        
        # Find the most recent revision before or on this date
        applicable_apr = base_apr
        for revision in self.interest_rate_revisions:
            if revision['date'].date() <= date.date():
                applicable_apr = revision['apr']
            else:
                break  # Since revisions are sorted, we can stop here
        
        return applicable_apr        
        
    def calculate(self):
        """Calculate loan amortization schedule"""
        try:
            # Get input values
            loan_amount = float(self.loan_amount.text())
            base_apr = float(self.apr.text())
            year_base = int(self.year_base.text())
            emi_amount = float(self.emi.text())
            emi_day = int(self.emi_date.text())
            tenure_months = int(self.loan_tenure.text())
            
            start_date = self.loan_start_dt.date()
            start_datetime = datetime(start_date.year(), start_date.month(), start_date.day())
            
            interest_date_value = self.interest_charged_date.currentText()
            
            # Initialize variables
            remaining_balance = loan_amount
            cumulative_interest = 0
            last_interest_debit_row = -1  # Track when last interest was debited
            total_interest_paid = 0
            
            # Clear existing table
            self.schedule_table.setRowCount(0)
            
            # Calculate end date based on tenure
            end_date = start_datetime + relativedelta(months=tenure_months)
            
            # Generate date-wise amortization schedule
            current_date = start_datetime
            row = 0
            emi_count = 0
            interest_debit_count = 0
            
            while current_date < end_date and remaining_balance > 0.01:
                # Add row to table
                self.schedule_table.insertRow(row)
                
                # Calculate Beginning Balance for this row
                if row == 0:
                    # First row - use original loan amount
                    beginning_balance = loan_amount
                else:
                    # From second row onwards: Previous Beginning Balance + Misc Charges + Interest Debited - EMI - Prepayment
                    prev_beginning = float(self.schedule_table.item(row-1, 1).text().replace('₹', '').replace(',', ''))
                    
                    prev_misc = 0
                    if self.schedule_table.item(row-1, 2) and self.schedule_table.item(row-1, 2).text():
                        prev_misc = float(self.schedule_table.item(row-1, 2).text().replace('₹', '').replace(',', ''))
                    
                    prev_interest_debited = 0
                    if self.schedule_table.item(row-1, 7) and self.schedule_table.item(row-1, 7).text():
                        prev_interest_debited = float(self.schedule_table.item(row-1, 7).text().replace('₹', '').replace(',', ''))
                    
                    prev_emi = 0
                    if self.schedule_table.item(row-1, 8) and self.schedule_table.item(row-1, 8).text():
                        prev_emi = float(self.schedule_table.item(row-1, 8).text().replace('₹', '').replace(',', ''))
                    
                    prev_prepayment = 0
                    if self.schedule_table.item(row-1, 9) and self.schedule_table.item(row-1, 9).text():
                        prev_prepayment = float(self.schedule_table.item(row-1, 9).text().replace('₹', '').replace(',', ''))
                    
                    beginning_balance = prev_beginning + prev_misc + prev_interest_debited - prev_emi - prev_prepayment
                
                remaining_balance = beginning_balance
                
                # Date
                self.set_table_item(row, 0, current_date.strftime("%d-%m-%Y"))
                
                # Beginning Balance
                self.set_table_item(row, 1, f"₹{beginning_balance:,.2f}")
                
                # Misc Charges By Bank - check for any bank charges on this date
                bank_charge = self.get_bank_charge_for_date(current_date)
                self.set_table_item(row, 2, f"₹{bank_charge:,.2f}" if bank_charge > 0 else "")
                
                # Get applicable APR for current date
                current_apr = self.get_apr_for_date(current_date)
                current_daily_rate = current_apr / (year_base * 100)
                
                # Interest Rate (Annual)
                self.set_table_item(row, 3, f"{current_apr}%")
                
                # Interest Rate (Daily)
                self.set_table_item(row, 4, f"{current_daily_rate*100:.6f}%")
                
                # Check if this is an EMI date
                is_emi_date = current_date.day == emi_day
                
                # Check if this month is excluded from automatic EMI
                is_excluded_month = self.is_emi_excluded(current_date)
                
                # Get manual EMI for this date
                manual_emi = self.get_manual_emi_for_date(current_date)
                
                # EMI (regular EMI on EMI date if not excluded + any manual EMI)
                emi_paid = 0
                if is_emi_date and not is_excluded_month:
                    emi_paid = emi_amount
                    emi_count += 1
                
                # Add manual EMI
                emi_paid += manual_emi
                
                self.set_table_item(row, 8, f"₹{emi_paid:,.2f}" if emi_paid > 0 else "")
                
                # Pre-Payment - check for any prepayments on this date
                prepayment = self.get_prepayment_for_date(current_date)
                self.set_table_item(row, 9, f"₹{prepayment:,.2f}" if prepayment > 0 else "")
                
                # Calculate adjusted balance for interest calculation
                # Beginning Balance + Misc Charges - EMI - Prepayment (EXCLUDING Interest Debited)
                adjusted_balance = beginning_balance + bank_charge - emi_paid - prepayment

                # Interest Amount for Each Day (calculated on adjusted balance, NOT including interest debited)
                daily_interest = adjusted_balance * current_daily_rate
                cumulative_interest += daily_interest
                self.set_table_item(row, 5, f"₹{daily_interest:,.2f}")
                
                # Check if this is an interest charged date
                is_interest_date = False
                if interest_date_value == "EOM":
                    # Check if this is the last day of the month
                    last_day = calendar.monthrange(current_date.year, current_date.month)[1]
                    is_interest_date = current_date.day == last_day
                else:
                    interest_day = int(interest_date_value)
                    is_interest_date = current_date.day == interest_day
                

                interest_debited = 0
                if is_interest_date:
                    # Sum all daily interest from last debit to current row (including current row)
                    start_row = last_interest_debit_row + 1
                    for sum_row in range(start_row, row + 1):  # INCLUDES current row
                        daily_int_item = self.schedule_table.item(sum_row, 5)  # Column 5 is "Interest Amount for Each Day"
                        if daily_int_item and daily_int_item.text():
                            daily_int_value = float(daily_int_item.text().replace('₹', '').replace(',', ''))
                            interest_debited += daily_int_value
                    
                    # Round off to nearest whole number
                    interest_debited = round(interest_debited, 0)
                    
                    if interest_debited > 0:
                        interest_debit_count += 1
                        # Update last debit row
                        last_interest_debit_row = row
                        
                self.set_table_item(row, 7, f"₹{interest_debited:,.0f}" if interest_debited > 0 else "")
                
                # Interest Paid and Principal Paid calculation
                total_payment = emi_paid + prepayment
                interest_paid = 0
                principal_paid = 0
                display_cumulative_interest = cumulative_interest  # Store for display before adjustment

                if total_payment > 0:
                    if total_payment >= cumulative_interest:
                        # Payment covers all accumulated interest and more
                        interest_paid = cumulative_interest
                        principal_paid = total_payment - interest_paid
                        # Reset cumulative interest to 0 for next row
                        cumulative_interest = 0
                    else:
                        # Payment is less than accumulated interest
                        interest_paid = total_payment
                        principal_paid = 0
                        # Reduce cumulative interest by the amount paid
                        cumulative_interest = cumulative_interest - interest_paid

                self.set_table_item(row, 10, f"₹{interest_paid:,.2f}" if interest_paid > 0 else "")
                self.set_table_item(row, 11, f"₹{principal_paid:,.2f}" if principal_paid > 0 else "")

                # Cumulative Interest Accrued (display value BEFORE payment adjustment)
                # Cumulative Interest Accrued (display value BEFORE payment adjustment)
                self.set_table_item(row, 6, f"₹{display_cumulative_interest:,.2f}")
                
                # Remaining Balance: Beginning Balance + Misc Charges + Interest Debited - EMI - Prepayment
                remaining_balance = beginning_balance + bank_charge + interest_debited - emi_paid - prepayment
                self.set_table_item(row, 12, f"₹{remaining_balance:,.2f}")
                
                # Balance + Interest Due (balance + cumulative interest after any payments)
                balance_plus_interest = remaining_balance + cumulative_interest
                self.set_table_item(row, 13, f"₹{balance_plus_interest:,.2f}")
                
                # Total Interest Paid (add current row's interest paid)
                if interest_paid > 0:
                    total_interest_paid += interest_paid
                self.set_table_item(row, 14, f"₹{total_interest_paid:,.2f}")
                
                # Color code rows
                if bank_charge > 0:
                    # Orange for bank charge dates
                    for col in range(15):
                        item = self.schedule_table.item(row, col)
                        if item:
                            item.setBackground(QColor("#ffe6cc"))
                elif prepayment > 0:
                    # Purple for prepayment dates
                    for col in range(15):
                        item = self.schedule_table.item(row, col)
                        if item:
                            item.setBackground(QColor("#f3e5f5"))
                elif manual_emi > 0 and not is_emi_date:
                    # Cyan for manual EMI dates (not on regular EMI date)
                    for col in range(15):
                        item = self.schedule_table.item(row, col)
                        if item:
                            item.setBackground(QColor("#e0f7fa"))
                elif is_emi_date and is_excluded_month:
                    # Light gray for excluded EMI dates
                    for col in range(15):
                        item = self.schedule_table.item(row, col)
                        if item:
                            item.setBackground(QColor("#f5f5f5"))
                elif is_interest_date and interest_debited > 0:
                    # Yellow for interest charged dates
                    for col in range(15):
                        item = self.schedule_table.item(row, col)
                        if item:
                            item.setBackground(QColor("#fff9e6"))
                elif is_emi_date and emi_paid > 0:
                    # Green for EMI payment rows
                    for col in range(15):
                        item = self.schedule_table.item(row, col)
                        if item:
                            item.setBackground(QColor("#e8f8f5"))
                
                # Move to next day
                current_date += timedelta(days=1)
                row += 1
                
                # Safety break for very long calculations
                if row > 10000:
                    break            # Update summary
            total_payment = emi_amount * emi_count
            total_prepayment = sum(self.get_prepayment_for_date(start_datetime + timedelta(days=d)) 
                                  for d in range(row))
            total_bank_charges = sum(charge['amount'] for charge in self.bank_charges)
            total_manual_emis = sum(emi['amount'] for emi in self.manual_emis)
            
            summary = f"""
═══════════════════════════════════════════════════════════════
                    LOAN CALCULATION SUMMARY
═══════════════════════════════════════════════════════════════

Loan Amount              : ₹{loan_amount:,.2f}
Initial Interest Rate    : {base_apr}%
Interest Rate Revisions  : {len(self.interest_rate_revisions)}
Daily Interest Rate      : Variable (based on revisions)
EMI Amount               : ₹{emi_amount:,.2f}
EMI Date (Each Month)    : {emi_day}
Excluded EMI Months      : {len(self.emi_exclusions)}
Interest Charged Date    : {interest_date_value} (each month)
Loan Tenure              : {tenure_months} months
Total Days               : {row} days
Regular EMI Payments     : {emi_count}
Interest Debits Made     : {interest_debit_count}
Pre-Payments Added       : {len(self.prepayments)}
Bank Charges Added       : {len(self.bank_charges)}
Manual EMIs Added        : {len(self.manual_emis)}

Total Regular EMI Paid   : ₹{total_payment:,.2f}
Total Manual EMI Paid    : ₹{total_manual_emis:,.2f}
Total Pre-Payments       : ₹{total_prepayment:,.2f}
Total Bank Charges       : ₹{total_bank_charges:,.2f}
Total Amount Paid        : ₹{total_payment + total_prepayment + total_manual_emis:,.2f}
Total Interest Paid      : ₹{total_interest_paid:,.2f}
Final Remaining Balance  : ₹{remaining_balance:,.2f}

Color Legend:
  🟠 Orange = Bank Charge Date
  🟪 Purple = Pre-Payment Date
  🔵 Cyan = Manual EMI Date
  ⚪ Gray = Excluded EMI Date
  🟨 Yellow = Interest Charged Date  
  🟩 Green = Regular EMI Payment Date

Note: Double-click any Pre-Payment cell to add/edit prepayment for that date
Note: Interest rates are revised based on the dates specified in revisions

═══════════════════════════════════════════════════════════════
            """
            self.summary_text.setText(summary)
            
            # Switch to schedule tab
            self.tab_widget.setCurrentIndex(1)
            
        except Exception as e:
            self.summary_text.setText(f"Error in calculation: {str(e)}\n\nPlease check your input values.") 

    def export_to_excel(self):
        """Export the amortization schedule to Excel"""
        try:
            if self.schedule_table.rowCount() == 0:
                # Show message if no data to export
                from PyQt6.QtWidgets import QMessageBox
                QMessageBox.warning(self, "No Data", "Please calculate the loan schedule first before exporting.")
                return
            
            # Create file dialog to get save location
            from PyQt6.QtWidgets import QFileDialog
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Save Excel File",
                "loan_schedule.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return  # User cancelled
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Amortization Schedule"
            
            # Define styles
            header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            # Colors for different row types
            bank_charge_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
            prepayment_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
            manual_emi_fill = PatternFill(start_color="E0F7FA", end_color="E0F7FA", fill_type="solid")
            excluded_emi_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            interest_date_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
            emi_fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
            
            center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            right_alignment = Alignment(horizontal="right", vertical="center")
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Add summary information at the top
            loan_amount = float(self.loan_amount.text())
            base_apr = float(self.apr.text())
            emi_amount = float(self.emi.text())
            tenure_months = int(self.loan_tenure.text())
            
            ws['A1'] = "LOAN CALCULATION SUMMARY"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:E1')
            
            summary_data = [
                ["Loan Amount:", f"₹{loan_amount:,.2f}"],
                ["Initial Interest Rate:", f"{base_apr}%"],
                ["EMI Amount:", f"₹{emi_amount:,.2f}"],
                ["Loan Tenure:", f"{tenure_months} months"],
                ["Total Rows:", str(self.schedule_table.rowCount())],
            ]
            
            row_num = 2
            for label, value in summary_data:
                ws[f'A{row_num}'] = label
                ws[f'A{row_num}'].font = Font(bold=True)
                ws[f'B{row_num}'] = value
                row_num += 1
            
            # Add blank row
            row_num += 1
            
            # Write headers
            headers = [
                "Date", "Beginning Balance", "Misc Charges By Bank", 
                "Interest Rate (Annual)", "Interest Rate (Daily)", 
                "Interest Amount for Each Day", "Cumulative Interest Accrued",
                "Interest Debited By Bank", "EMI", "Pre-Payment",
                "Interest Paid", "Principal Paid", "Remaining Balance",
                "Balance + Interest Due", "Total Interest Paid"
            ]
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border
            
            header_row = row_num
            row_num += 1
            
            # Write data rows
            for table_row in range(self.schedule_table.rowCount()):
                # Get background color from table
                bg_color = self.schedule_table.item(table_row, 0).background().color()
                
                # Determine fill based on color
                fill = None
                if bg_color.name() == "#ffe6cc":
                    fill = bank_charge_fill
                elif bg_color.name() == "#f3e5f5":
                    fill = prepayment_fill
                elif bg_color.name() == "#e0f7fa":
                    fill = manual_emi_fill
                elif bg_color.name() == "#f5f5f5":
                    fill = excluded_emi_fill
                elif bg_color.name() == "#fff9e6":
                    fill = interest_date_fill
                elif bg_color.name() == "#e8f8f5":
                    fill = emi_fill
                
                for col_num in range(15):
                    item = self.schedule_table.item(table_row, col_num)
                    cell = ws.cell(row=row_num, column=col_num + 1)
                    
                    if item and item.text():
                        value = item.text()
                        
                        # Remove currency symbol and commas for numeric columns
                        if col_num == 0:  # Date column
                            cell.value = value
                            cell.alignment = center_alignment
                        elif col_num in [3, 4]:  # Percentage columns
                            cell.value = value
                            cell.alignment = center_alignment
                        else:  # Numeric columns
                            try:
                                numeric_value = float(value.replace('₹', '').replace(',', '').replace('%', ''))
                                cell.value = numeric_value
                                if col_num in [3, 4]:  # Interest rate columns
                                    cell.number_format = '0.000000"%"'
                                else:
                                    cell.number_format = '₹#,##0.00'
                                cell.alignment = right_alignment
                            except:
                                cell.value = value
                                cell.alignment = center_alignment
                    
                    cell.border = thin_border
                    if fill:
                        cell.fill = fill
                
                row_num += 1
            
            # Auto-adjust column widths
            for col_num in range(1, 16):
                column_letter = get_column_letter(col_num)
                max_length = 0
                
                # Check header length
                if ws[f'{column_letter}{header_row}'].value:
                    max_length = len(str(ws[f'{column_letter}{header_row}'].value))
                
                # Check data length (sample first 100 rows for performance)
                for row in range(header_row + 1, min(header_row + 101, row_num)):
                    cell_value = ws[f'{column_letter}{row}'].value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                
                adjusted_width = min(max_length + 2, 30)  # Cap at 30
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze panes (freeze header row and first column)
            ws.freeze_panes = f'B{header_row + 1}'
            
            # Add legend at the bottom
            row_num += 2
            ws[f'A{row_num}'] = "COLOR LEGEND:"
            ws[f'A{row_num}'].font = Font(bold=True)
            row_num += 1
            
            legend_data = [
                ("Bank Charge Date", bank_charge_fill),
                ("Pre-Payment Date", prepayment_fill),
                ("Manual EMI Date", manual_emi_fill),
                ("Excluded EMI Date", excluded_emi_fill),
                ("Interest Charged Date", interest_date_fill),
                ("Regular EMI Payment Date", emi_fill),
            ]
            
            for label, fill_style in legend_data:
                ws[f'A{row_num}'] = label
                ws[f'A{row_num}'].fill = fill_style
                ws[f'A{row_num}'].border = thin_border
                row_num += 1
            
            # Save workbook
            wb.save(file_path)
            
            # Show success message
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.information(self, "Success", f"Excel file saved successfully!\n\n{file_path}")
            
        except Exception as e:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.critical(self, "Error", f"Failed to export to Excel:\n{str(e)}")

            
    def set_table_item(self, row, col, value):
        """Helper method to set table item"""
        item = QTableWidgetItem(value)
        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.schedule_table.setItem(row, col, item)
    
    def last_day_of_month(self, date):
        """Get last day of month"""
        next_month = date.replace(day=28) + timedelta(days=4)
        return (next_month - timedelta(days=next_month.day)).day
    
    def clear_fields(self):
        """Clear all input fields"""
        self.loan_amount.clear()
        self.loan_start_dt.setDate(QDate.currentDate())
        self.interest_charged_date.setCurrentIndex(0)
        self.loan_tenure.clear()
        self.apr.clear()
        self.emi.clear()
        self.year_base.clear()
        self.emi_date.clear()
        self.prepayments = []
        self.bank_charges = []
        self.manual_emis = []
        self.emi_exclusions = []
        self.view_prepayments_btn.setText("View (0)")
        self.view_bank_charges_btn.setText("View (0)")
        self.view_manual_emis_btn.setText("View (0)")
        self.exclude_emi_btn.setText("Ex (0)")
        self.summary_text.clear()
        self.schedule_table.setRowCount(0)

    def get_settings_file_path(self):
        """Get the path to the settings file"""
        app_data_dir = os.path.expanduser("~/.loan_calculator")
        if not os.path.exists(app_data_dir):
            os.makedirs(app_data_dir)
        return os.path.join(app_data_dir, "settings.json")
    
    def save_settings(self):
        """Save all input fields to a JSON file"""
        try:
            settings = {
                'loan_amount': self.loan_amount.text(),
                'apr': self.apr.text(),
                'year_base': self.year_base.text(),
                'loan_start_date': self.loan_start_dt.date().toString("dd-MM-yyyy"),
                'emi': self.emi.text(),
                'emi_date': self.emi_date.text(),
                'interest_charged_date': self.interest_charged_date.currentText(),
                'loan_tenure': self.loan_tenure.text(),
                'prepayments': [
                    {
                        'type': pp['type'],
                        'amount': pp['amount'],
                        'date': pp.get('date').strftime('%d-%m-%Y') if pp.get('date') else None,
                        'day': pp.get('day'),
                        'start_date': pp.get('start_date').strftime('%d-%m-%Y') if pp.get('start_date') else None,
                        'end_date': pp.get('end_date').strftime('%d-%m-%Y') if pp.get('end_date') else None
                    }
                    for pp in self.prepayments
                ],
                'bank_charges': [
                    {
                        'amount': bc['amount'],
                        'date': bc['date'].strftime('%d-%m-%Y'),
                        'description': bc['description']
                    }
                    for bc in self.bank_charges
                ],
                'manual_emis': [
                    {
                        'amount': me['amount'],
                        'date': me['date'].strftime('%d-%m-%Y'),
                        'note': me['note']
                    }
                    for me in self.manual_emis
                ],
                'emi_exclusions': self.emi_exclusions,
                'interest_rate_revisions': [
                    {
                        'apr': rev['apr'],
                        'date': rev['date'].strftime('%d-%m-%Y')
                    }
                    for rev in self.interest_rate_revisions
                ]
            }
            
            with open(self.get_settings_file_path(), 'w') as f:
                json.dump(settings, f, indent=4)
        except Exception as e:
            print(f"Error saving settings: {e}")
    
    def load_settings(self):
        """Load all input fields from JSON file"""
        try:
            settings_file = self.get_settings_file_path()
            if not os.path.exists(settings_file):
                return
            
            with open(settings_file, 'r') as f:
                settings = json.load(f)
            
            # Load basic fields
            self.loan_amount.setText(settings.get('loan_amount', ''))
            self.apr.setText(settings.get('apr', ''))
            self.year_base.setText(settings.get('year_base', ''))
            self.emi.setText(settings.get('emi', ''))
            self.emi_date.setText(settings.get('emi_date', ''))
            self.loan_tenure.setText(settings.get('loan_tenure', ''))
            
            # Load loan start date
            date_str = settings.get('loan_start_date')
            if date_str:
                self.loan_start_dt.setDate(QDate.fromString(date_str, "dd-MM-yyyy"))
            
            # Load interest charged date
            interest_date = settings.get('interest_charged_date')
            if interest_date:
                self.interest_charged_date.setCurrentText(interest_date)
            
            # Load prepayments
            self.prepayments = []
            for pp in settings.get('prepayments', []):
                prepayment = {'type': pp['type'], 'amount': pp['amount']}
                if pp['type'] == 'single':
                    prepayment['date'] = datetime.strptime(pp['date'], '%d-%m-%Y')
                elif pp['type'] == 'recurring':
                    prepayment['day'] = pp['day']
                    prepayment['start_date'] = datetime.strptime(pp['start_date'], '%d-%m-%Y')
                    prepayment['end_date'] = datetime.strptime(pp['end_date'], '%d-%m-%Y') if pp['end_date'] else None
                self.prepayments.append(prepayment)
            self.view_prepayments_btn.setText(f"View ({len(self.prepayments)})")
            
            # Load bank charges
            self.bank_charges = []
            for bc in settings.get('bank_charges', []):
                self.bank_charges.append({
                    'amount': bc['amount'],
                    'date': datetime.strptime(bc['date'], '%d-%m-%Y'),
                    'description': bc['description']
                })
            self.view_bank_charges_btn.setText(f"View ({len(self.bank_charges)})")
            
            # Load manual EMIs
            self.manual_emis = []
            for me in settings.get('manual_emis', []):
                self.manual_emis.append({
                    'amount': me['amount'],
                    'date': datetime.strptime(me['date'], '%d-%m-%Y'),
                    'note': me['note']
                })
            self.view_manual_emis_btn.setText(f"View ({len(self.manual_emis)})")
            
            # Load EMI exclusions
            self.emi_exclusions = settings.get('emi_exclusions', [])
            self.exclude_emi_btn.setText(f"Ex ({len(self.emi_exclusions)})")
            
            # Load interest rate revisions
            self.interest_rate_revisions = []
            for rev in settings.get('interest_rate_revisions', []):
                self.interest_rate_revisions.append({
                    'apr': rev['apr'],
                    'date': datetime.strptime(rev['date'], '%d-%m-%Y')
                })
            # Sort by date
            self.interest_rate_revisions.sort(key=lambda x: x['date'])
            self.view_rate_revisions_btn.setText(f"View ({len(self.interest_rate_revisions)})")
            
        except Exception as e:
            print(f"Error loading settings: {e}")
    
    def closeEvent(self, event):
        """Override close event to save settings"""
        self.save_settings()
        event.accept()

def main():
    app = QApplication(sys.argv)
    font = QFont("Segoe UI", 10)
    app.setFont(font)
    window = LoanCalculatorApp()
    window.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()